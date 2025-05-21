import logging
import os
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation  # Import InvalidOperation

from ofxstatement.parser import StatementParser
from ofxstatement.plugin import Plugin
from ofxstatement.statement import Statement, StatementLine, generate_transaction_id
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger("SantanderES")

"""
Data wrapper
"""


@dataclass
class Movement:
    operation_data: datetime
    value_data: datetime
    description: str
    amount: Decimal  # Keep type hint as Decimal
    balance: Decimal  # Keep type hint as Decimal
    currency: str
    stat_line: StatementLine = None

    def __post_init__(self):

        logging.debug(
            f"Movement(operation_data={self.operation_data}, value_data={self.value_data}, description={self.description}, amount={self.amount}, balance={self.balance}, currency={self.currency})"
        )

        # Convert operation_data to datetime if it's a string.
        if isinstance(self.operation_data, str):
            try:
                self.operation_data = datetime.strptime(self.operation_data, "%d/%m/%Y")
            except ValueError:
                raise ValueError(
                    "operation_data must be a datetime or a string in format DD/MM/YYYY"
                )

        # Convert value_data to datetime if it's a string.
        if isinstance(self.value_data, str):
            try:
                self.value_data = datetime.strptime(self.value_data, "%d/%m/%Y")
            except ValueError:
                raise ValueError(
                    "value_data must be a datetime or a string in format DD/MM/YYYY"
                )

        # Convert amount to Decimal if it's a string
        if isinstance(self.amount, str):
            try:
                # Clean the string: replace the specific minus sign '−' with '-',
                # remove thousands separators (.), replace decimal comma (,) with dot (.)
                amount_str_cleaned = (
                    self.amount.replace("−", "-").replace(".", "").replace(",", ".")
                )
                logging.debug(
                    f"Cleaned amount string: '{amount_str_cleaned}'"
                )  # Add debug log here
                self.amount = Decimal(amount_str_cleaned)
            except InvalidOperation:
                raise ValueError(f"Could not convert amount '{self.amount}' to Decimal")

        # Convert balance to Decimal if it's a string
        if isinstance(self.balance, str):
            try:
                # Clean the string: replace the specific minus sign '−' with '-',
                # remove thousands separators (.), replace decimal comma (,) with dot (.)
                balance_str_cleaned = (
                    self.balance.replace("−", "-").replace(".", "").replace(",", ".")
                )
                logging.debug(
                    f"Cleaned balance string: '{balance_str_cleaned}'"
                )  # Add debug log here
                self.balance = Decimal(balance_str_cleaned)
            except InvalidOperation:
                # Balance might not always be critical for StatementLine, but good to convert
                logging.warning(
                    f"Could not convert balance '{self.balance}' to Decimal"
                )
                # Optionally set balance to None or 0 if conversion fails
                self.balance = Decimal(0)  # Or None, depending on desired behavior

        # Now create the StatementLine with a proper datetime for the date and Decimal for amount
        self.stat_line = StatementLine(
            None,
            self.operation_data,
            self.description,
            self.amount,  # This is now a Decimal
        )
        # Generate and assign a unique transaction ID
        self.stat_line.id = generate_transaction_id(self.stat_line)
        # Optionally set the user date
        self.stat_line.date_user = self.value_data


class SantanderPlugin(Plugin):
    """SantanderEs plugin for ofxstatement"""

    def get_parser(self, filename):
        parser = SantanderXlsxParser(filename, self.settings)
        return parser


class SantanderXlsxParser(StatementParser):
    wb = None
    settings = None
    tableStart: int = 5

    def __init__(self, filename, settings):
        self.file = filename
        self.wb = load_workbook(self.file)
        self.settings = settings
        logging.debug(self.settings)

        if "Sheet1" in self.wb.sheetnames:
            logging.debug('Detected "Sheet1", using excel parser')
            while True:
                val = self.wb["Sheet1"][f"A{self.tableStart}"].value
                self.tableStart += 1
                if val == "Fecha operación":
                    break
                elif self.tableStart == 100:
                    logging.error(
                        'Unknown excel format, no "Fecha operación" found, aborting'
                    )
                    exit(os.EX_IOERR)
        else:
            logging.error("Unknown excel format, aborting")
            exit(os.EX_IOERR)

        self.statement = Statement(currency="EUR")
        logging.debug(self.statement)

    """
    Override method, use to obrain iterable object consisting of a line per
    transaction
    """

    def split_records(self) -> Iterator[Movement]:  # Corrected return type hint
        return self._get_movements()

    """
    Override method, use to generate Statement object
    """

    def parse(self):
        # The base parse method calls split_records and then parse_record for each item
        # Ensure the statement object has the correct balance and dates set after parsing
        statement = super(SantanderXlsxParser, self).parse()

        # After parsing all records, set the start/end balance and dates on the statement
        # This assumes the last row's balance is the start balance and the first row's date is the end date
        if statement.lines:
            # Set end balance to the balance of the first row
            statement.end_balance = Decimal(
                str(statement.lines[0].balance)
            )  # Ensure Decimal type

            # Calculate the sum of amounts using Decimal(0) as the start value
            total_amount = sum(
                (Decimal(line.amount) for line in statement.lines), Decimal(0)
            )

            # Calculate start balance based on the new end balance
            calculated_start_balance = statement.end_balance - total_amount
            statement.start_balance = Decimal(
                str(calculated_start_balance)
            )  # Convert via string for safety

            # Set end date (first row's date) and start date (last row's date)
            statement.end_date = statement.lines[0].date
            statement.start_date = statement.lines[-1].date

        return statement

    """
    Override use to Parse given transaction line and return StatementLine
    object
    """

    def parse_record(self, mov: Movement) -> StatementLine:
        logging.debug(f"Parsing movement: {mov}")
        # The StatementLine is already created and populated in Movement.__post_init__
        # We just need to ensure the balance is also set on the StatementLine if needed by ofxstatement
        # ofxstatement typically uses the balance from the Statement object, not individual lines,
        # but setting it here doesn't hurt and might be useful for debugging or other plugins.
        if mov.stat_line:
            mov.stat_line.balance = mov.balance  # Set balance on the StatementLine

        return mov.stat_line

    """ Private method to parse all record lines """

    def _get_movements(self) -> Iterator[Movement]:
        starting_column = column_index_from_string("A")
        ending_column = column_index_from_string("F")  # Assuming balance is in column F
        starting_row = self.tableStart

        for row in self.wb["Sheet1"].iter_rows(
            starting_row, None, starting_column, ending_column, values_only=True
        ):
            # Check if the first column (Fecha operación) is not empty
            if row[0] is not None:
                logging.debug(f"Processing row: {row}")
                try:
                    # Pass the raw row data to Movement. __post_init__ will handle conversions
                    yield Movement(*row)
                except ValueError as e:
                    logging.error(f"Error processing row {row}: {e}")
                    # Decide how to handle rows that fail conversion - skip, log, etc.
                    # For now, we re-raise the error as per the __post_init__ logic
                    raise e
            else:
                # Stop iterating when the first column is empty, assuming end of data
                logging.debug("Empty cell in first column, stopping iteration.")
                break
